import openpyxl
wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.get_sheet_by_name('Sheet')

price_updates_dict = {'Garlic':3.07, 'Lemon': 1.27}

for rowNum in range(2, sheet.max_row, 1):
    produceName = sheet.cell(rowNum, 1).value
    if produceName in price_updates_dict:
        sheet.cell(rowNum, 2).value = price_updates_dict[produceName]

wb.save('produceSales_update.xlsx')